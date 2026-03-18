#!/usr/bin/env python3
# multiprocess_prototype\database\export_detections.py
"""
Экспорт детекций из inspector.db в читаемый формат.

Использование:
  python -m multiprocess_prototype.database.export_detections [--format txt|csv|xlsx] [--output FILE]

  Из каталога Inspector_prototype:
    python -m multiprocess_prototype.database.export_detections
    python -m multiprocess_prototype.database.export_detections --format csv --output detections.csv
    python -m multiprocess_prototype.database.export_detections --format xlsx --output detections.xlsx
"""
import argparse
import csv
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

# Путь к БД относительно этого скрипта
_DB_PATH = Path(__file__).resolve().parent / "inspector.db"
_COLUMNS = ["id", "timestamp", "frame_name", "frame_id", "x1", "y1", "x2", "y2", "center_x", "center_y", "area"]


def _timestamp_to_str(ts: float) -> str:
    """Преобразовать Unix timestamp в читаемую строку."""
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    except (ValueError, OSError):
        return str(ts)


def export_txt(db_path: Path, output_path: Path) -> None:
    """Экспорт в текстовый файл (читаемый формат)."""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.execute("SELECT * FROM detections ORDER BY id")
    rows = cur.fetchall()
    conn.close()

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=" * 100 + "\n")
        f.write("Детекции из inspector.db\n")
        f.write(f"Экспорт: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Записей: {len(rows)}\n")
        f.write("=" * 100 + "\n\n")

        for row in rows:
            d = dict(row)
            ts_str = _timestamp_to_str(d.get("timestamp", 0))
            f.write(f"ID: {d.get('id')}\n")
            f.write(f"  Время:     {ts_str}\n")
            f.write(f"  Кадр:      {d.get('frame_name')} (id={d.get('frame_id')})\n")
            f.write(f"  Bbox:      ({d.get('x1')}, {d.get('y1')}) - ({d.get('x2')}, {d.get('y2')})\n")
            f.write(f"  Центр:     ({d.get('center_x')}, {d.get('center_y')})\n")
            f.write(f"  Площадь:   {d.get('area')} px\n")
            f.write("-" * 50 + "\n")
        if not rows:
            f.write("(нет записей)\n")

    print(f"Экспорт TXT: {output_path} ({len(rows)} записей)")


def export_csv(db_path: Path, output_path: Path) -> None:
    """Экспорт в CSV (открывается в Excel)."""
    conn = sqlite3.connect(db_path)
    cur = conn.execute("SELECT * FROM detections ORDER BY id")
    rows = cur.fetchall()
    conn.close()

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(_COLUMNS)
        for row in rows:
            writer.writerow(row)

    print(f"Экспорт CSV: {output_path} ({len(rows)} записей)")


def export_xlsx(db_path: Path, output_path: Path) -> None:
    """Экспорт в Excel (.xlsx). Требует openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Для экспорта в Excel установите: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    conn = sqlite3.connect(db_path)
    cur = conn.execute("SELECT * FROM detections ORDER BY id")
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"

    for col, name in enumerate(_COLUMNS, 1):
        ws.cell(row=1, column=col, value=name)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Автоширина колонок
    for col in range(1, len(_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(output_path)
    print(f"Экспорт XLSX: {output_path} ({len(rows)} записей)")


def main() -> int:
    parser = argparse.ArgumentParser(description="Экспорт детекций из inspector.db")
    parser.add_argument(
        "--format", "-f",
        choices=["txt", "csv", "xlsx"],
        default="txt",
        help="Формат вывода (по умолчанию: txt)",
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        help="Путь к выходному файлу (по умолчанию: detections.<format>)",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=_DB_PATH,
        help=f"Путь к БД (по умолчанию: {_DB_PATH})",
    )
    args = parser.parse_args()

    if not args.db.exists():
        print(f"Ошибка: файл БД не найден: {args.db}", file=sys.stderr)
        return 1

    output = args.output or Path(f"detections.{args.format}")
    output = output.resolve()

    if args.format == "txt":
        export_txt(args.db, output)
    elif args.format == "csv":
        export_csv(args.db, output)
    else:
        export_xlsx(args.db, output)

    return 0


if __name__ == "__main__":
    sys.exit(main())
