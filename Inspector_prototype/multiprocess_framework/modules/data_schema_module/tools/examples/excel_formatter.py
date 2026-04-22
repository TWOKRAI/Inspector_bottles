"""
Пример расширения: Excel форматер для визуализации схем.

Этот файл демонстрирует, как легко добавить новый формат визуализации
через паттерн Strategy.

Для использования требуется установить openpyxl:
    pip install openpyxl
"""

from typing import Dict, Any
from pathlib import Path
from io import BytesIO

from ...core.interfaces import IVisualizationFormatter


class ExcelVisualizationFormatter(IVisualizationFormatter):
    """
    Excel форматер визуализации схем.
    
    Экспортирует информацию о схеме в Excel файл.
    
    Example:
        from data_schema_module.extensions.tools import SchemaVisualizer
        from data_schema_module.tools.examples.excel_formatter import (
            ExcelVisualizationFormatter
        )
        
        visualizer = SchemaVisualizer()
        visualizer.register_formatter(ExcelVisualizationFormatter())
        
        # Сохранение в Excel
        visualizer.save_visualization(
            "MySchema",
            Path("schema.xlsx"),
            format="excel"
        )
    """
    
    @property
    def format_name(self) -> str:
        return "excel"
    
    def format(self, schema_name: str, schema_info: Dict[str, Any]) -> bytes:
        """
        Форматировать информацию о схеме в Excel формат.
        
        Args:
            schema_name: Имя схемы
            schema_info: Словарь с информацией о схеме
            
        Returns:
            Байты Excel файла
            
        Note:
            Этот метод возвращает bytes, а не str, что демонстрирует гибкость интерфейса.
            SchemaVisualizer.save_visualization() обрабатывает это корректно.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "Для использования Excel форматера требуется openpyxl. "
                "Установите: pip install openpyxl"
            )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Schema Info"
        
        # Заголовок
        ws['A1'] = f"Схема: {schema_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # Описание
        if schema_info.get("docstring"):
            ws['A2'] = "Описание:"
            ws['A2'].font = Font(bold=True)
            ws['B2'] = schema_info["docstring"]
            ws.merge_cells('B2:E2')
            ws['B2'].alignment = Alignment(wrap_text=True)
        
        # Заголовки таблицы
        headers = ["Поле", "Тип", "Обязательное", "По умолчанию", "Описание"]
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Данные полей
        row = 5
        for field in schema_info.get("fields", []):
            ws.cell(row=row, column=1).value = field["name"]
            ws.cell(row=row, column=2).value = field.get("type", "Any")
            ws.cell(row=row, column=3).value = "Да" if field.get("required", True) else "Нет"
            ws.cell(row=row, column=4).value = field.get("default", "-")
            ws.cell(row=row, column=5).value = field.get("description", "-")
            row += 1
        
        # Автоматическая ширина колонок
        for col in range(1, 6):
            max_length = 0
            column = ws.column_dimensions[chr(64 + col)]  # A, B, C, D, E
            for cell in ws[chr(64 + col)]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            column.width = min(max_length + 2, 50)
        
        # Сохраняем в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# Пример использования с базой данных (концептуальный)
class DatabaseVisualizationFormatter(IVisualizationFormatter):
    """
    Пример форматера для сохранения в базу данных.
    
    Демонстрирует концепцию расширения для работы с БД.
    """
    
    def __init__(self, connection_string: str):
        """
        Args:
            connection_string: Строка подключения к БД
        """
        self.connection_string = connection_string
    
    @property
    def format_name(self) -> str:
        return "database"
    
    def format(self, schema_name: str, schema_info: Dict[str, Any]) -> str:
        """
        Сохранить информацию о схеме в базу данных.
        
        Returns:
            ID записи в БД
        """
        # Здесь была бы реальная реализация сохранения в БД
        # Например:
        # import sqlalchemy
        # engine = sqlalchemy.create_engine(self.connection_string)
        # with engine.connect() as conn:
        #     result = conn.execute(
        #         sqlalchemy.text("INSERT INTO schemas ..."),
        #         {"name": schema_name, "info": json.dumps(schema_info)}
        #     )
        #     return str(result.lastrowid)
        
        return f"Saved to DB: {schema_name}"

