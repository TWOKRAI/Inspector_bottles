# -*- coding: utf-8 -*-
"""
TableExporter — экспорт List[Dict] в txt (читаемый/таблица), csv, xlsx.

Работает с объектом данных (rows), не с файлом напрямую.
Методы to_* возвращают строку или bytes; save() записывает в файл.
"""
from __future__ import annotations

import csv
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple, Union


class ExportFormat(str, Enum):
    """Форматы экспорта."""

    TXT_READABLE = "txt_readable"  # Читаемый формат (ID, поля построчно)
    TXT_TABLE = "txt_table"        # Обычная таблица с разделителем
    CSV = "csv"                    # CSV для Excel (;)
    XLSX = "xlsx"                  # Excel (.xlsx)


def _default_timestamp_formatter(v: Any) -> str:
    """Форматирование Unix timestamp в читаемую строку."""
    if v is None:
        return ""
    try:
        return datetime.fromtimestamp(float(v)).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    except (ValueError, OSError, TypeError):
        return str(v)


class TableExporter:
    """
    Экспорт табличных данных (List[Dict]) в файлы.

    Поддерживает:
    - TXT_READABLE: читаемый формат (каждая запись — блок с полями)
    - TXT_TABLE: таблица с разделителем |
    - CSV: разделитель ; (Excel)
    - XLSX: Excel (требует openpyxl)
    """

    def __init__(
        self,
        columns: Optional[List[str]] = None,
        formatters: Optional[Dict[str, Callable[[Any], str]]] = None,
        readable_labels: Optional[Dict[str, str]] = None,
        readable_blocks: Optional[List[Tuple[str, Callable[[Dict], str]]]] = None,
    ):
        """
        Args:
            columns: Порядок колонок. None — из первой строки.
            formatters: Функции форматирования по имени колонки (для TXT_READABLE).
            readable_labels: Человекочитаемые названия полей (для TXT_READABLE).
            readable_blocks: Для TXT_READABLE — составные поля (label, getter). Если заданы,
                используются вместо columns для читаемого формата.
        """
        self.columns = columns
        self.formatters = formatters or {}
        self.readable_labels = readable_labels or {}
        self.readable_blocks = readable_blocks
        self.formatters.setdefault("timestamp", _default_timestamp_formatter)

    def _get_columns(self, rows: List[Dict[str, Any]]) -> List[str]:
        """Получить список колонок."""
        if self.columns:
            return self.columns
        if rows:
            return list(rows[0].keys())
        return []

    def _format_value(self, col: str, value: Any) -> str:
        """Форматировать значение для вывода."""
        fn = self.formatters.get(col)
        if fn:
            return fn(value)
        return "" if value is None else str(value)

    def to_txt_readable(
        self,
        rows: List[Dict[str, Any]],
        title: str = "Экспорт данных",
        export_time: Optional[datetime] = None,
    ) -> str:
        """Читаемый формат: каждая запись — блок с полями."""
        lines = []
        lines.append("=" * 100)
        lines.append(title)
        lines.append(f"Экспорт: {(export_time or datetime.now()).strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Записей: {len(rows)}")
        lines.append("=" * 100)
        lines.append("")

        if self.readable_blocks:
            for row in rows:
                for label, getter in self.readable_blocks:
                    lines.append(f"{label}: {getter(row)}")
                lines.append("-" * 50)
        else:
            cols = self._get_columns(rows)
            for row in rows:
                for col in cols:
                    label = self.readable_labels.get(col, col)
                    val = self._format_value(col, row.get(col))
                    lines.append(f"{label}: {val}")
                lines.append("-" * 50)

        if not rows:
            lines.append("(нет записей)")

        return "\n".join(lines)

    def to_txt_table(
        self,
        rows: List[Dict[str, Any]],
        delimiter: str = " | ",
        header: bool = True,
    ) -> str:
        """Таблица с разделителем (обычная таблица)."""
        cols = self._get_columns(rows)
        lines = []

        if header and cols:
            lines.append(delimiter.join(str(c) for c in cols))
            lines.append(delimiter.join("-" * max(4, len(str(c))) for c in cols))

        for row in rows:
            parts = [self._format_value(col, row.get(col)) for col in cols]
            lines.append(delimiter.join(parts))

        return "\n".join(lines)

    def to_csv(
        self,
        rows: List[Dict[str, Any]],
        delimiter: str = ";",
    ) -> str:
        """CSV-строка (разделитель ; для Excel)."""
        cols = self._get_columns(rows)
        import io
        buf = io.StringIO()
        writer = csv.writer(buf, delimiter=delimiter)
        writer.writerow(cols)
        for row in rows:
            writer.writerow([row.get(c) for c in cols])
        return buf.getvalue()

    def to_xlsx_bytes(self, rows: List[Dict[str, Any]], sheet_name: str = "Data") -> bytes:
        """Excel как bytes. Требует openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise ImportError("Для экспорта в Excel установите: pip install openpyxl") from e

        cols = self._get_columns(rows)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        for col_idx, name in enumerate(cols, 1):
            ws.cell(row=1, column=col_idx, value=name)

        for row_idx, row in enumerate(rows, 2):
            for col_idx, col in enumerate(cols, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col))

        for col in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def save(
        self,
        rows: List[Dict[str, Any]],
        path: Union[str, Path],
        format: Union[ExportFormat, str] = ExportFormat.TXT_READABLE,
        **kwargs: Any,
    ) -> int:
        """
        Сохранить rows в файл.

        Args:
            rows: Данные (List[Dict])
            path: Путь к файлу
            format: TXT_READABLE, TXT_TABLE, CSV, XLSX
            **kwargs: title, delimiter, sheet_name и т.д.

        Returns:
            Количество записанных записей.
        """
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        fmt = ExportFormat(format) if isinstance(format, str) else format

        if fmt == ExportFormat.TXT_READABLE:
            content = self.to_txt_readable(
                rows,
                title=kwargs.get("title", "Экспорт данных"),
                export_time=kwargs.get("export_time"),
            )
            path.write_text(content, encoding="utf-8")
        elif fmt == ExportFormat.TXT_TABLE:
            content = self.to_txt_table(
                rows,
                delimiter=kwargs.get("delimiter", " | "),
                header=kwargs.get("header", True),
            )
            path.write_text(content, encoding="utf-8")
        elif fmt == ExportFormat.CSV:
            content = self.to_csv(rows, delimiter=kwargs.get("delimiter", ";"))
            path.write_text(content, encoding="utf-8-sig")
        elif fmt == ExportFormat.XLSX:
            data = self.to_xlsx_bytes(rows, sheet_name=kwargs.get("sheet_name", "Data"))
            path.write_bytes(data)
        else:
            raise ValueError(f"Неизвестный формат: {format}")

        return len(rows)
